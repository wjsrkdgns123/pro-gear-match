import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import urllib.parse

mice = [
  "Logitech G Pro X Superlight 2","Logitech G Pro X Superlight","Logitech G Pro Wireless","Logitech G703 Lightspeed","Logitech G502 X Plus","Logitech G502 Hero","Logitech G403 Hero","Logitech G305 Lightspeed","Logitech G102 Lightsync",
  "Razer DeathAdder V3 Pro","Razer Viper V3 Pro","Razer Viper V2 Pro","Razer Viper Mini Signature Edition","Razer DeathAdder V2 Pro","Razer DeathAdder V2","Razer Basilisk V3 Pro","Razer Cobra Pro","Razer Orochi V2","Razer Viper Ultimate",
  "Zowie EC2-CW","Zowie EC1-CW","Zowie EC3-CW","Zowie EC2-B","Zowie EC2-C","Zowie FK2-C","Zowie FK1+-C","Zowie S2-C","Zowie S1-C","Zowie ZA13-C","Zowie ZA12-C","Zowie ZA11-C",
  "VAXEE XE Wireless","VAXEE NP-01 Wireless","VAXEE NP-01S Wireless","VAXEE OUTSET AX Wireless","VAXEE ZYGEN NP-01",
  "Finalmouse UltralightX","Finalmouse Starlight-12","Finalmouse Starlight Pro - TenZ","Finalmouse Air58 Ninja","Finalmouse Ultralight 2",
  "Pulsar X2V2","Pulsar Xlite V3","Pulsar X2H","Pulsar X2A","Pulsar Xlite V2 Wireless",
  "Lamzu Atlantis OG V2","Lamzu Maya","Lamzu Thorn","Lamzu Energon",
  "SteelSeries Prime Wireless","SteelSeries Prime Mini Wireless","SteelSeries Aerox 3 Wireless","SteelSeries Aerox 5 Wireless","SteelSeries Rival 3","SteelSeries Rival 650",
  "HyperX Pulsefire Haste 2","HyperX Pulsefire Dart",
  "G-Wolves HTS Plus 4K","G-Wolves Hati-S Plus","G-Wolves HSK Pro 4K",
  "Scyrox V8","Ninjutso Sora V2","Ninjutso Katana Superlight","Ninjutso Origin One X",
  "Endgame Gear XM2we","Endgame Gear OP1 8k","Endgame Gear OP1we","Endgame Gear XM1r",
  "Glorious Model O 2 Wireless","Glorious Model D- Wireless","Glorious Model I 2 Wireless",
  "Xtrfy M4 Wireless","Xtrfy M42 Wireless","Xtrfy MZ1 Wireless","Xtrfy M8 Wireless",
  "Roccat Kone Pro Air","Roccat Burst Pro","Cooler Master MM711","Cooler Master MM712",
  "Asus ROG Harpe Ace Aim Lab Edition","Asus ROG Keris Wireless AimPoint",
  "Cherry Xtrfy M64 Pro","Cherry Xtrfy M68 Pro","VGN Dragonfly F1 Pro Max","VXE R1 Pro","Zaopin Z1 Pro",
  "BenQ Zowie FK1-B","BenQ Zowie S2 Divina","TRUST GXT 179 Gamers",
  "SteelSeries Rival 5","SteelSeries Rival 600","SteelSeries Sensei Ten",
  "Corsair Dark Core RGB Pro","Corsair M65 RGB Ultra","Corsair Ironclaw RGB",
  "Mad Catz R.A.T. 8+ Pro","Mad Catz R.A.T. Pro X","Turtle Beach Krait Ultra",
  "Gigabyte Aorus M5","MSI Clutch GM11","Turtle Beach Krait","Kingston Fury Beast Pro","Patriot Viper V560",
  "ZOWIE FK2-DW Black","ZOWIE EC2-DW Black","ZOWIE EC2-DW Glossy","ZOWIE EC2-DW Grey","ZOWIE EC3-DW Black","ZOWIE EC3-DW Glossy","ZOWIE U2-DW Black","ZOWIE FK1-C","ZOWIE x donk Mouse",
  "Finalmouse Ultralight X Medium","Finalmouse Ultralight X Large","Finalmouse Ultralight X Prophecy ScreaM Classic","Finalmouse Ultralight X Prophecy Clix Small",
  "HITSCAN Hyperlight Black","Pulsar X2 V2 Black","Pulsar X2N","Pulsar X2N Crazylight Medium","Pulsar Xlite V4 Es Medium","Pulsar eS FS-1",
  "Pulsar ZywOo The Chosen Mouse Black","Pulsar ZywOo The Chosen Mouse White","Pulsar TenZ Signature Edition","PULSAR TENZ EDITION","Pulsar Susanto-X",
  "Logitech G303 Shroud Edition","Logitech G903","Logitech G Pro X2 SUPERSTRIKE","Logitech G Pro 2 LIGHTSPEED Magenta",
  "G-Wolves Hati S+ 4K","G-Wolves Hati S+ 8K",
  "Razer Viper V4 Pro","Razer Viper V4 Pro Black","Razer Viper V3 Hyperspeed","Razer DeathAdder V3 HyperSpeed",
  "Razer DeathAdder V4 Pro Black","Razer DeathAdder V4 Pro White","Razer DeathAdder V4 Pro Green","Razer DeathAdder V4 Pro NiKo Edition","Razer Basilisk Ultimate",
  "SONY INZONE Mouse-A","VGN Dragonfly MOBA","Alienware Pro Wireless Gaming Mouse","Lamzu Atlantis Mini",
  "Endgame Gear OP1w 4K Black","ROCCAT Kain 200 Aimo","WLMouse Beast X Max Red","WLMouse BEAST X Silver","Fallen Gear Lobo Wireless",
  "Corsair Sabre v2 Pro","VAXEE XE V2 White","VAXEE XE V2 Black","VAXEE XE V2 Pink","VAXEE XE V2 Fluorescent Green","VAXEE E1 Wireless White","VAXEE E1 Wireless Red",
  "Cooler Master MM710","ATK Blazing Sky F1","ATK Dragonfly A9","ASUS ROG Harpe Ace 2 Black","Corsair Katar Pro Wireless"
]

keyboards = [
  "Wooting 60HE","Wooting 80HE","Wooting Two HE",
  "SteelSeries Apex Pro TKL (2023)","SteelSeries Apex Pro Mini","SteelSeries Apex Pro TKL Wireless","SteelSeries Apex 9 TKL","SteelSeries Apex 7 TKL",
  "Razer Huntsman V3 Pro TKL","Razer Huntsman V3 Pro Mini","Razer Huntsman V3 Pro Full Size","Razer Huntsman V2 TKL","Razer BlackWidow V4 Pro","Razer BlackWidow V3 TKL",
  "Logitech G Pro X TKL Wireless","Logitech G Pro X Keyboard","Logitech G915 TKL","Logitech G715","Logitech G815","Logitech G613",
  "Varmilo VA87M","Varmilo MA87M V2",
  "Ducky One 3 TKL","Ducky One 2 Mini","Ducky Mecha Mini","Ducky SF",
  "HyperX Alloy FPS Pro","HyperX Alloy Origins Core","HyperX Alloy Elite 2",
  "Corsair K70 RGB TKL","Corsair K65 RGB Mini","Corsair K100 RGB","Corsair K70 RGB MK.2",
  "Custom Mechanical Keyboard","Matrix Premium","Higround Base 65","Higround Basecamp 65","Higround Summit 65",
  "Keychron Q1","Keychron Q2","Keychron K2","Keychron V1","Keychron K6",
  "DrunkDeer A75","DrunkDeer G65","Leopold FC750R","Leopold FC660M","Leopold FC900R",
  "Filco Majestouch 2","Realforce R2 TKL","Realforce GX1","HHKB Professional Hybrid Type-S",
  "Anne Pro 2","Akko 3068","Akko Mod 007 v3","GMMK Pro","GMMK 2","Glorious GMMK TKL",
  "ZSA Moonlander","NuPhy Air75","IQUNIX L80","Mistel Barocco MD770","Monsgeek King Kong 87",
  "Epomaker GK96S","Ajazz AK820","Gateron Apex Pro",
  "Corsair K70 Lux","Steelseries Apex M800","Razer Huntsman Elite","Razer Pro Type Ultra","GMMK 2 Pro",
  "Keychron K8 Pro","iKBC CD108","Novelkeys NK87 Entry","Input Club K-Type",
  "Meletrix BOOG75","Razer BlackWidow V3 Pro","Razer BlackWidow V3 Mini","Razer BlackWidow V4 75%","Razer Huntsman Mini","Razer Huntsman TE","Razer Blackwidow Chroma TE V2",
  "Higround Performance Base 65","Durgod Taurus K320","HyperX Alloy Origins 60",
  "ASUS ROG Falchion RX Low Profile","ASUS ROG Falchion Ace HFX Black","ASUS ROG Falchion Ace HFX ZywOo Edition","ASUS ROG Falchion Ace 75 HE White","ASUS ROG Falchion Ace 75 HE Black","ASUS ROG Falcata",
  "Logitech G Pro Keyboard","Logitech G Pro Mechanical Gaming Keyboard","Logitech G Pro X TKL Keyboard","Logitech G Pro X TKL RAPID","Logitech G Pro X TKL RAPID Black","Logitech G Pro X TKL RAPID White","Logitech G Pro X 60 Black","Logitech G512","Logitech G610","Logitech G713","Logitech G913 TKL",
  "ROCCAT Vulcan TKL Pro","Corsair K70 PRO TKL Black",
  "SteelSeries Apex Pro TKL Gen 3 White","SteelSeries Apex Pro TKL Gen 3 Black",
  "Pulsar PCMK 2 HE TKL White","Pulsar PCMK 3","Fnatic miniSTREAK","Sony INZONE KBD-H75","ZENAIM Keyboard",
  "ATK RS6 Ultra Aspas Edition","Wooting 60 HE","Logitech G Pro X Mechanical Keyboard","Logitech G Pro Mechanical Keyboard","Logitech G Pro X Mechanical Gaming Keyboard"
]

monitors = [
  "Zowie XL2566K (360Hz)","Zowie XL2546K (240Hz)","Zowie XL2546 (240Hz)","Zowie XL2540 (240Hz)","Zowie XL2746K (240Hz)","Zowie XL2546S (240Hz)","Zowie XL2586X (540Hz)","Zowie XL2546X (240Hz)","Zowie XL2540K (240Hz)","Zowie XL2411K (144Hz)",
  "ASUS ROG Swift PG259QN (360Hz)","ASUS ROG Swift PG27AQN (360Hz)","ASUS ROG Swift PG258Q (240Hz)","ASUS ROG Swift Pro PG248QP (540Hz)","ASUS ROG Swift PG27AQDM (240Hz OLED)","ASUS ROG Swift PG259QNR (360Hz)","ASUS TUF Gaming VG259QM (280Hz)","ASUS TUF Gaming VG249Q (144Hz)","ASUS VG248QE (144Hz)",
  "Alienware AW2521H (360Hz)","Alienware AW2523HF (360Hz)","Alienware AW2518H (240Hz)","Alienware AW2524H (500Hz)","Alienware AW2723DF (280Hz)",
  "LG UltraGear 27GR95QE (240Hz OLED)","LG UltraGear 27GN750 (240Hz)","LG UltraGear 24GN600 (144Hz)","LG UltraGear 27GP850 (165Hz)",
  "Samsung Odyssey G7 (240Hz)","Samsung Odyssey G9 (240Hz)","Samsung Odyssey G6 (240Hz)","Samsung Odyssey G4 (240Hz)",
  "MSI G251FG (360Hz)","MSI Optix MAG251RX (240Hz)","MSI MAG 271QPX (240Hz OLED)",
  "Acer Predator X25 (360Hz)","Acer Nitro XV252Q F (390Hz)",
  "ViewSonic XG2431 (240Hz)","ViewSonic XG251G (360Hz)",
  "AOC 24G2 (144Hz)","AOC AGON AG251FZ (240Hz)","AOC C24G1 (144Hz)",
  "HP Omen X 25 (240Hz)","Gigabyte M27Q (170Hz)","Gigabyte G24F (165Hz)",
  "ZOWIE XL2566X+","ZOWIE XL2586X+","BenQ ZOWIE XL2566K","BenQ ZOWIE XL2546K","BenQ ZOWIE XL2546",
  "SONY INZONE M10S","AOC C27G2ZE","LG ULTRAGEAR 27GN950-B","LG ULTRAGEAR 32GQ950-B","LG ULTRAGEAR 27GN750-B",
  "MSI Optix G271","Dell S2522HG","ASUS TUF VG259QM","ASUS TUF VG258QM","ASUS ROG XG258Q","ASUS ROG STRIX XG248Q",
  "Acer Predator XB252Q","Acer Predator XB272",
  "Alienware AW2721D","Alienware AW2521HF","Alienware AW3225QF",
  "Samsung C24RG54FQU","HP OMEN 27i","HP OMEN 24.5","OMEN 25i","OMEN 25","OMEN by HP 25"
]

mousepads = [
  "Artisan Zero (Soft)","Artisan Ninja FX Zero Soft Black","Artisan Zero (XSoft)","Artisan Zero (Mid)","Artisan Hayate Otsu (Soft)","Artisan Hayate Otsu (XSoft)","Artisan Hayate Otsu (Mid)","Artisan Hien (Soft)","Artisan Hien (Mid)","Artisan Hien (XSoft)","Artisan Raiden (XSoft)","Artisan Raiden (Mid)","Artisan Shidenkai V2","Artisan Kou (Soft)",
  "SteelSeries Qck Heavy","SteelSeries Qck+","SteelSeries Qck Edge","SteelSeries Qck Mass","SteelSeries Qck Prism",
  "Zowie G-SR-SE (Gris/Rouge)","Zowie G-SR-SE (Deep Blue)","Zowie G-SR-SE (Bi)","Zowie G-SR-SE (Tyloo)","Zowie G-SR-SE (Pink/Divina)","Zowie G-SR II","Zowie G-SR (Black)","Zowie P-SR",
  "Logitech G640","Logitech G840","Logitech G740","Logitech G240","Logitech G440 (Hard)",
  "Razer Gigantus V2","Razer Strider","Razer Goliathus Speed","Razer Goliathus Control","Razer Firefly V2",
  "VAXEE PA","VAXEE PB",
  "Lethal Gaming Gear Saturn Pro","Lethal Gaming Gear Venus Pro","Lethal Gaming Gear Saturn","Lethal Gaming Gear Venus","Lethal Gaming Gear Jupiter",
  "SkyPAD Glass 3.0","Wallhack SP-004","Wallhack MO-001","BenQ Zowie G-SR",
  "HyperX Fury S Pro","HyperX Fury S Speed","Xtrfy GP4","Xtrfy GP1","Xtrfy GP2",
  "Fnatic Dash","Fnatic Focus 3","Fnatic Jet","Gamesense Radar",
  "Aqua Control 2","Aqua Control+","Aqua Control 2 (Sakura)","X-raypad Equate Plus",
  "The Mousepad Company Strata","Odin Gaming ZeroGravity","Pulsar ES2","Pulsar ES1","Pulsar Paraspeed","Pulsar Superglide (Glass)",
  "PureTrak Talent","Glorious Fire","Glorious Ice","Corsair MM300","Corsair MM350","Cooler Master MP511","Endgame Gear OP1 8K","Razer Strider Pro","Cooler Master MP511 XL",
  "Ducky Sherlock Dark","Ducky Sherlock Frost","Logitech G Pro X","SteelSeries QcK Large","SteelSeries QcK XL","SteelSeries QcK Performance Speed",
  "HyperX Fury L Pro","Corsair MM800","Corsair MM500","Turtle Beach Flash Pro","Vanguard VG Infinity","Glorious 3XL Extended","Razer Invicta","Mad Catz Glide 38","Kingston Fury Beast XL","Patriot Viper V1 Pro",
  "Artisan Ninja FX Zero Soft Orange","Artisan Ninja FX Zero XSoft Black","Artisan Ninja FX Zero XSoft Orange","Artisan Ninja FX Zero XSoft","Artisan Ninja FX Zero Mid","Artisan Ninja FX Zero Mid Orange","Artisan Ninja FX Shidenkai XSoft Black","Artisan Hayate Otsu Soft Black","Artisan Hayate Otsu Soft Wine Red","Artisan Hayate Otsu XSoft Black","Artisan Hayate Otsu Mid Wine Red","Artisan FX Hien Soft Black","Artisan FX Hien XSoft Wine Red","Artisan Raiden FX Mid","Artisan Raiden FX XSoft","Artisan Raiden FX Soft Coffee Brown","Artisan Key-83 XSoft","Artisan Type-99 Soft Black","Artisan Type-99 XSoft Black","Artisan Type-99 Gray","Artisan Type-99 Matcha",
  "SkyPAD 3.0","SkyPAD 3.0 XL",
  "ZOWIE G-SR-SE Rouge","ZOWIE G-SR-SE ROUGE II","ZOWIE G-SR-SE Blue","ZOWIE G-SR-SE BLUE II","ZOWIE G-SR-SE Bi","ZOWIE G-SR-SE Bi II","ZOWIE G-SR-SE Divina Blue","ZOWIE G-SR-SE Gris","ZOWIE G-TR","ZOWIE H-SR III","ZOWIE H-SR-SE BLUE II",
  "Pulsar x LGG Saturn Pro Black","Pulsar x LGG Venus Pro Red","Pulsar x LGG Jupiter Pro","Pulsar eS Saturn Pro Black","Pulsar eS Saturn Pro Red",
  "Xraypad Aqua Control II Black","Higround Performance Mousepad","Yuki Aim Oni Black","Yuki Aim x Demon1 Cloth","InfinityMice Infinite Speed Matcha Green","Fallen Gear Mystic",
  "Fnatic FOCUS 2","Fnatic FOCUS3 MAX","Abyss Gaming Scarlet","Roccat Sense AIMO","A4Tech X7-500MP","CHERRY XTRFY GP7","Degster Victorious Dawn","Odin Gaming Andromeda Pro","BanKs Collection Heavy Claw by ESPTIGER",
  "VAXEE PD150","VAXEE PD151","VAXEE PE Kumo","VAXEE PE-K Pink",
  "SONY INZONE MAT-D","La Onda QCW Molten Edge","The Chosen Mousepad","Mionix Sargas","Artisan Ninja FX Zero (Soft)"
]

wb = openpyxl.Workbook()

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
data_font = Font(name='Arial', size=10)
link_font = Font(name='Arial', size=10, color='1565C0', underline='single')
alt_fill = PatternFill('solid', start_color='F1F8E9')

categories = [
    ('마우스 (Mice)', mice, '1565C0'),
    ('키보드 (Keyboards)', keyboards, '6A1B9A'),
    ('모니터 (Monitors)', monitors, 'B71C1C'),
    ('마우스패드 (Mousepads)', mousepads, '1B5E20'),
]

# Summary sheet
ws_sum = wb.active
ws_sum.title = '요약'
ws_sum.column_dimensions['A'].width = 32
ws_sum.column_dimensions['B'].width = 15

ws_sum['A1'] = 'Pro Gear Match - 장비 전체 목록'
ws_sum['A1'].font = Font(name='Arial', bold=True, size=14)
ws_sum.merge_cells('A1:B1')
ws_sum.row_dimensions[1].height = 28

ws_sum.append([])
ws_sum.append(['카테고리', '장비 수'])
for col in range(1, 3):
    ws_sum.cell(3, col).font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    ws_sum.cell(3, col).fill = PatternFill('solid', start_color='37474F')
    ws_sum.cell(3, col).alignment = center
    ws_sum.cell(3, col).border = border

for i, (cat, items, _) in enumerate(categories):
    r = 4 + i
    ws_sum.cell(r, 1, cat).font = data_font
    ws_sum.cell(r, 2, len(items)).font = data_font
    ws_sum.cell(r, 1).alignment = left_align
    ws_sum.cell(r, 2).alignment = center
    for col in range(1, 3):
        ws_sum.cell(r, col).border = border
    if i % 2 == 0:
        for col in range(1, 3):
            ws_sum.cell(r, col).fill = PatternFill('solid', start_color='ECEFF1')

total_row = 8
ws_sum.cell(total_row, 1, '합계').font = Font(name='Arial', bold=True, size=11)
ws_sum.cell(total_row, 2, '=SUM(B4:B7)').font = Font(name='Arial', bold=True, size=11)
ws_sum.cell(total_row, 1).alignment = center
ws_sum.cell(total_row, 2).alignment = center
for col in range(1, 3):
    ws_sum.cell(total_row, col).border = border
    ws_sum.cell(total_row, col).fill = PatternFill('solid', start_color='B2DFDB')

# Category sheets
for cat, items, color in categories:
    short_name = cat.split(' ')[0]
    ws = wb.create_sheet(short_name)

    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 58

    hf = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill('solid', start_color=color)

    headers = ['#', '제품명', '아마존 검색 링크']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(1, col_idx, h).font = hf
        ws.cell(1, col_idx, h).fill = hfill
        ws.cell(1, col_idx, h).alignment = center
        ws.cell(1, col_idx, h).border = border
    ws.row_dimensions[1].height = 22

    for i, name in enumerate(items, start=1):
        row = i + 1
        query = urllib.parse.quote_plus(name)
        url = "https://www.amazon.com/s?k=" + query

        ws.cell(row, 1, i).font = data_font
        ws.cell(row, 1).alignment = center
        ws.cell(row, 2, name).font = data_font
        ws.cell(row, 2).alignment = left_align
        ws.cell(row, 3).hyperlink = url
        ws.cell(row, 3).value = "Amazon: " + name
        ws.cell(row, 3).font = link_font
        ws.cell(row, 3).alignment = left_align

        for col in range(1, 4):
            ws.cell(row, col).border = border
            if i % 2 == 0:
                ws.cell(row, col).fill = alt_fill
        ws.row_dimensions[row].height = 17

output_path = 'C:/Users/wjsrk/Desktop/ProGearMatch_Equipment_List.xlsx'
wb.save(output_path)
print("Saved:", output_path)
print("Mice:", len(mice), "| Keyboards:", len(keyboards), "| Monitors:", len(monitors), "| Mousepads:", len(mousepads))
print("Total:", len(mice) + len(keyboards) + len(monitors) + len(mousepads))
